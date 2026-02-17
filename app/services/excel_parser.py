"""
Excel Parser Service - Parse Paytm and other Excel exports
"""
import hashlib
import uuid
from datetime import datetime
from typing import List, Dict, Optional, Tuple, BinaryIO
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

# Try to import openpyxl, provide helpful error if not installed
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Install with: pip install openpyxl")


class ParsedTransaction:
    """Represents a single parsed transaction from Excel"""
    def __init__(
        self,
        row_number: int,
        date: Optional[str] = None,
        time: Optional[str] = None,
        description: Optional[str] = None,
        amount: Optional[float] = None,
        transaction_type: str = "Expense",
        category: Optional[str] = None,
        subcategory: Optional[str] = None,
        account: Optional[str] = None,
        status: Optional[str] = None,
        reference_id: Optional[str] = None,
        is_valid: bool = True,
        error_message: Optional[str] = None
    ):
        self.row_number = row_number
        self.date = date
        self.time = time
        self.description = description
        self.amount = amount
        self.transaction_type = transaction_type
        self.category = category
        self.subcategory = subcategory
        self.account = account
        self.status = status
        self.reference_id = reference_id
        self.is_valid = is_valid
        self.error_message = error_message

    def to_dict(self) -> dict:
        return {
            "row_number": self.row_number,
            "date": self.date,
            "time": self.time,
            "description": self.description,
            "amount": self.amount,
            "transaction_type": self.transaction_type,
            "category": self.category,
            "subcategory": self.subcategory,
            "account": self.account,
            "status": self.status,
            "reference_id": self.reference_id,
            "is_valid": self.is_valid,
            "error_message": self.error_message
        }


class ColumnMappingInfo:
    """Information about a column mapping"""
    def __init__(self, excel_column: str, internal_field: str, sample_value: Optional[str] = None):
        self.excel_column = excel_column
        self.internal_field = internal_field
        self.sample_value = sample_value
    
    def to_dict(self) -> dict:
        return {
            "excel_column": self.excel_column,
            "internal_field": self.internal_field,
            "sample_value": self.sample_value
        }


class ParseResult:
    """Result of parsing an Excel file"""
    def __init__(
        self,
        filename: str,
        file_hash: str,
        file_type: str,
        transactions: List[ParsedTransaction],
        total_rows: int,
        valid_count: int,
        skipped_count: int,
        column_mappings: List[ColumnMappingInfo] = None,
        unique_accounts: List[str] = None
    ):
        self.filename = filename
        self.file_hash = file_hash
        self.file_type = file_type
        self.transactions = transactions
        self.total_rows = total_rows
        self.valid_count = valid_count
        self.skipped_count = skipped_count
        self.preview_token = str(uuid.uuid4())
        self.column_mappings = column_mappings or []
        self.unique_accounts = unique_accounts or []

    def to_dict(self) -> dict:
        return {
            "filename": self.filename,
            "file_hash": self.file_hash,
            "file_type": self.file_type,
            "total_rows": self.total_rows,
            "valid_transactions": self.valid_count,
            "skipped_rows": self.skipped_count,
            "preview_token": self.preview_token,
            "transactions": [t.to_dict() for t in self.transactions],
            "column_mappings": [m.to_dict() for m in self.column_mappings],
            "unique_accounts": self.unique_accounts
        }


def get_file_hash(content: bytes) -> str:
    """Calculate SHA256 hash of file content"""
    return hashlib.sha256(content).hexdigest()


def parse_paytm_excel(
    file_content: bytes,
    filename: str
) -> ParseResult:
    """
    Parse Paytm export Excel file.
    
    Paytm UPI Statement format (multiple sheets):
    - Sheet "Summary": Account summary (skipped)
    - Sheet "Passbook Payment History": Actual transactions
    
    Columns in Passbook Payment History:
    - Date, Time, Transaction Details, Other Transaction Details, Your Account, Amount, UPI Ref No., Order ID, Remarks, Tags, Comment
    
    Args:
        file_content: Raw bytes of the Excel file
        filename: Original filename
        
    Returns:
        ParseResult with parsed transactions
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")
    
    file_hash = get_file_hash(file_content)
    transactions: List[ParsedTransaction] = []
    
    # Load workbook from bytes
    wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    
    # Try to find the transaction sheet (Paytm uses "Passbook Payment History")
    ws = None
    transaction_sheet_names = ["passbook payment history", "transactions", "history", "statement"]
    
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in transaction_sheet_names or "passbook" in sheet_name.lower() or "history" in sheet_name.lower():
            ws = wb[sheet_name]
            logger.info(f"Found transaction sheet: {sheet_name}")
            break
    
    # Fallback to active sheet if no transaction sheet found
    if ws is None:
        ws = wb.active
        logger.info(f"Using active sheet: {ws.title}")
    
    # Find header row and map columns
    header_row = None
    column_map = {}
    
    # Common Paytm column name variations
    date_columns = ["date", "transaction date", "txn date"]
    time_columns = ["time", "txn time", "transaction time"]
    description_columns = ["activity", "description", "narration", "details", "particulars", "transaction details"]
    amount_columns = ["amount", "value", "sum"]  # Single column with +/- sign
    debit_columns = ["debit", "dr", "amount debited", "paid"]
    credit_columns = ["credit", "cr", "amount credited", "received"]
    status_columns = ["status", "txn status", "transaction status"]
    reference_columns = ["transaction id", "txn id", "reference", "ref no", "reference no", "upi ref no.", "upi ref no"]
    tags_columns = ["tags", "category", "tag"]
    account_columns = ["your account", "account", "from account", "source account"]
    
    # Find header row (usually row 1 or 2)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        row_values = [str(cell).lower().strip() if cell else "" for cell in row]
        
        # Check if this looks like a header row
        has_date = any(col in row_values for col in date_columns)
        has_amount = any(col in row_values for col in amount_columns + debit_columns + credit_columns)
        
        if has_date and has_amount:
            header_row = row_idx
            # Map column indices
            for idx, cell_value in enumerate(row_values):
                if cell_value in date_columns:
                    column_map["date"] = idx
                elif cell_value in time_columns:
                    column_map["time"] = idx
                elif cell_value in description_columns:
                    column_map["description"] = idx
                elif cell_value in amount_columns:
                    column_map["amount"] = idx  # Single amount column with +/- sign
                elif cell_value in debit_columns:
                    column_map["debit"] = idx
                elif cell_value in credit_columns:
                    column_map["credit"] = idx
                elif cell_value in status_columns:
                    column_map["status"] = idx
                elif cell_value in reference_columns:
                    column_map["reference"] = idx
                elif cell_value in tags_columns:
                    column_map["tags"] = idx
                elif cell_value in account_columns:
                    column_map["account"] = idx
            break
    
    if header_row is None:
        # Couldn't find header, assume first row with Paytm UPI format
        header_row = 1
        column_map = {
            "date": 0,
            "time": 1,
            "description": 2,
            "amount": 5,  # Column F in Paytm format
            "reference": 6
        }
    
    logger.info(f"Found header at row {header_row}, column map: {column_map}")
    
    # Build column mapping info for preview
    internal_field_names = {
        "date": "Date",
        "time": "Time",
        "description": "Notes / Vendor",
        "amount": "Amount",
        "debit": "Amount (Debit)",
        "credit": "Amount (Credit)",
        "status": "Status",
        "reference": "Reference ID",
        "tags": "Category / Subcategory",
        "account": "Account"
    }
    
    # Get original header names for mapping display
    original_headers = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        for idx, cell in enumerate(row):
            if cell:
                original_headers[idx] = str(cell)
        break
    
    column_mappings = []
    for field, idx in column_map.items():
        excel_col = original_headers.get(idx, f"Column {idx+1}")
        internal_name = internal_field_names.get(field, field.title())
        column_mappings.append(ColumnMappingInfo(
            excel_column=excel_col,
            internal_field=internal_name,
            sample_value=None  # Will be populated with first row data
        ))
    
    total_rows = 0
    valid_count = 0
    skipped_count = 0
    unique_accounts_set = set()
    
    # Helper to safely get column value
    def get_col(row, key, default_idx=None):
        idx = column_map.get(key, default_idx)
        if idx is not None and len(row) > idx:
            return row[idx]
        return None
    
    # Parse data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        total_rows += 1
        
        try:
            # Extract values using column map
            date_val = get_col(row, "date", 0)
            time_val = get_col(row, "time")  # Separate time column
            description = get_col(row, "description", 2)
            amount_val = get_col(row, "amount")  # Single amount column with +/- sign
            debit = get_col(row, "debit")
            credit = get_col(row, "credit")
            status = get_col(row, "status")
            reference = get_col(row, "reference", 6)
            tags = get_col(row, "tags")
            account = get_col(row, "account")
            
            # Skip empty rows
            if not date_val and not description and not amount_val and not debit and not credit:
                continue
            
            # Parse date
            parsed_date = None
            parsed_time = None
            
            if date_val:
                if isinstance(date_val, datetime):
                    parsed_date = date_val.strftime("%Y-%m-%d")
                    parsed_time = date_val.strftime("%H:%M:%S")
                else:
                    # Try to parse string date (Paytm uses DD/MM/YYYY format)
                    date_str = str(date_val).strip()
                    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d-%b-%Y", "%d %b'%y", "%d %b '%y"]:
                        try:
                            dt = datetime.strptime(date_str.split()[0], fmt)
                            parsed_date = dt.strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            continue
            
            # Parse time from separate column if available
            if time_val and not parsed_time:
                if isinstance(time_val, datetime):
                    parsed_time = time_val.strftime("%H:%M:%S")
                else:
                    parsed_time = str(time_val).strip()
            
            # Determine amount and transaction type
            amount = None
            transaction_type = "Expense"
            
            # First check if there's a single "amount" column with +/- sign (Paytm UPI format)
            if amount_val:
                amount_str = str(amount_val).replace(",", "").replace("₹", "").replace(" ", "").strip()
                if amount_str and amount_str not in ["", "-", "0", "0.0", "0.00"]:
                    try:
                        # Check for + or - prefix to determine type
                        if amount_str.startswith("+"):
                            amount = abs(float(amount_str))
                            transaction_type = "Income"
                        elif amount_str.startswith("-"):
                            amount = abs(float(amount_str))
                            transaction_type = "Expense"
                        else:
                            # No sign - assume expense
                            amount = abs(float(amount_str))
                            transaction_type = "Expense"
                    except ValueError:
                        pass
            
            # Fallback to separate debit/credit columns
            if amount is None:
                if debit and str(debit).strip() not in ["", "0", "0.0", "-"]:
                    try:
                        amount = abs(float(str(debit).replace(",", "").replace("₹", "").strip()))
                        transaction_type = "Expense"
                    except ValueError:
                        pass
                
                if credit and str(credit).strip() not in ["", "0", "0.0", "-"]:
                    try:
                        amount = abs(float(str(credit).replace(",", "").replace("₹", "").strip()))
                        transaction_type = "Income"
                    except ValueError:
                        pass
            
            # Extract category and subcategory from tags if available (Paytm uses #Category: Subcategory format)
            category = None
            subcategory = None
            if tags:
                tags_str = str(tags).strip()
                # Extract category from tags like "#Food & Drinks: Groceries" or "#💵 Self-Transfer"
                if tags_str.startswith("#"):
                    tags_content = tags_str[1:].strip()
                    
                    # Remove emoji prefix if present (emoji + space)
                    if len(tags_content) > 2 and not tags_content[0].isalnum():
                        # Skip first character (emoji) and any following space
                        if len(tags_content) > 1 and tags_content[1] == ' ':
                            tags_content = tags_content[2:].strip()
                        elif len(tags_content) > 1:
                            tags_content = tags_content[1:].strip()
                    
                    # Split by colon to get category and subcategory
                    if ":" in tags_content:
                        parts = tags_content.split(":", 1)
                        category = parts[0].strip()
                        subcategory = parts[1].strip() if len(parts) > 1 else None
                    else:
                        category = tags_content.strip()
            
            # Build description with account info if available
            desc_str = str(description).strip() if description else ""
            
            # Check status - only import successful transactions
            # Note: Paytm doesn't have a separate status column in transaction list
            status_str = str(status).upper().strip() if status else "SUCCESS"
            is_valid = status_str in ["SUCCESS", "COMPLETED", "SUCCESSFUL", "", "NONE"]
            error_msg = None
            
            if not is_valid:
                error_msg = f"Transaction status: {status_str}"
                skipped_count += 1
            elif amount is None or amount == 0:
                is_valid = False
                error_msg = "No amount found"
                skipped_count += 1
            elif not parsed_date:
                is_valid = False
                error_msg = "Could not parse date"
                skipped_count += 1
            else:
                valid_count += 1
            
            # Get account name from "Your Account" column
            account_name = str(account).strip() if account else None
            
            # Track unique accounts
            if account_name and account_name.lower() != "none":
                unique_accounts_set.add(account_name)
            
            transactions.append(ParsedTransaction(
                row_number=row_idx,
                date=parsed_date,
                time=parsed_time,
                description=desc_str,
                amount=amount,
                transaction_type=transaction_type,
                category=category,
                subcategory=subcategory,
                account=account_name,
                status=status_str,
                reference_id=str(reference).strip() if reference else None,
                is_valid=is_valid,
                error_message=error_msg
            ))
            
        except Exception as e:
            logger.error(f"Error parsing row {row_idx}: {e}")
            skipped_count += 1
            transactions.append(ParsedTransaction(
                row_number=row_idx,
                is_valid=False,
                error_message=str(e)
            ))
    
    wb.close()
    
    # Add sample values to column mappings from first valid transaction
    first_valid = next((t for t in transactions if t.is_valid), None)
    if first_valid:
        for mapping in column_mappings:
            if mapping.internal_field == "Date":
                mapping.sample_value = first_valid.date
            elif mapping.internal_field == "Time":
                mapping.sample_value = first_valid.time
            elif mapping.internal_field == "Notes / Vendor":
                mapping.sample_value = first_valid.description[:50] if first_valid.description else None
            elif mapping.internal_field == "Amount":
                mapping.sample_value = str(first_valid.amount) if first_valid.amount else None
            elif mapping.internal_field == "Category / Subcategory":
                cat_sub = first_valid.category or ""
                if first_valid.subcategory:
                    cat_sub += f": {first_valid.subcategory}"
                mapping.sample_value = cat_sub if cat_sub else None
            elif mapping.internal_field == "Account":
                mapping.sample_value = first_valid.account
            elif mapping.internal_field == "Reference ID":
                mapping.sample_value = first_valid.reference_id
    
    return ParseResult(
        filename=filename,
        file_hash=file_hash,
        file_type="paytm_excel",
        transactions=transactions,
        total_rows=total_rows,
        valid_count=valid_count,
        skipped_count=skipped_count,
        column_mappings=column_mappings,
        unique_accounts=sorted(list(unique_accounts_set))
    )


def parse_generic_excel(
    file_content: bytes,
    filename: str,
    column_mapping: Optional[Dict[str, int]] = None
) -> ParseResult:
    """
    Parse a generic Excel file with custom column mapping.
    
    Args:
        file_content: Raw bytes of the Excel file
        filename: Original filename
        column_mapping: Dict mapping field names to column indices
            e.g., {"date": 0, "description": 1, "amount": 2}
    
    Returns:
        ParseResult with parsed transactions
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")
    
    file_hash = get_file_hash(file_content)
    transactions: List[ParsedTransaction] = []
    
    # Default mapping
    if column_mapping is None:
        column_mapping = {
            "date": 0,
            "description": 1,
            "amount": 2,
            "type": 3  # Expense/Income or Debit/Credit
        }
    
    wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    
    total_rows = 0
    valid_count = 0
    skipped_count = 0
    
    # Skip header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        total_rows += 1
        
        try:
            date_val = row[column_mapping.get("date", 0)] if len(row) > column_mapping.get("date", 0) else None
            description = row[column_mapping.get("description", 1)] if len(row) > column_mapping.get("description", 1) else None
            amount_val = row[column_mapping.get("amount", 2)] if len(row) > column_mapping.get("amount", 2) else None
            type_val = row[column_mapping.get("type")] if column_mapping.get("type") and len(row) > column_mapping.get("type") else None
            
            # Skip empty rows
            if not date_val and not description and not amount_val:
                continue
            
            # Parse date
            parsed_date = None
            if date_val:
                if isinstance(date_val, datetime):
                    parsed_date = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val).strip()
                    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"]:
                        try:
                            dt = datetime.strptime(date_str, fmt)
                            parsed_date = dt.strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            continue
            
            # Parse amount
            amount = None
            if amount_val:
                try:
                    amount = abs(float(str(amount_val).replace(",", "").replace("₹", "").strip()))
                except ValueError:
                    pass
            
            # Determine transaction type
            transaction_type = "Expense"
            if type_val:
                type_str = str(type_val).upper().strip()
                if type_str in ["INCOME", "CREDIT", "CR", "RECEIVED"]:
                    transaction_type = "Income"
            
            is_valid = amount is not None and amount > 0 and parsed_date is not None
            error_msg = None
            
            if not is_valid:
                error_msg = "Missing date or amount"
                skipped_count += 1
            else:
                valid_count += 1
            
            transactions.append(ParsedTransaction(
                row_number=row_idx,
                date=parsed_date,
                description=str(description).strip() if description else None,
                amount=amount,
                transaction_type=transaction_type,
                is_valid=is_valid,
                error_message=error_msg
            ))
            
        except Exception as e:
            logger.error(f"Error parsing row {row_idx}: {e}")
            skipped_count += 1
            transactions.append(ParsedTransaction(
                row_number=row_idx,
                is_valid=False,
                error_message=str(e)
            ))
    
    wb.close()
    
    return ParseResult(
        filename=filename,
        file_hash=file_hash,
        file_type="generic_excel",
        transactions=transactions,
        total_rows=total_rows,
        valid_count=valid_count,
        skipped_count=skipped_count
    )


def generate_csv_content(transactions: List[ParsedTransaction], include_invalid: bool = False) -> str:
    """
    Generate CSV content from parsed transactions.
    
    Format matches SpendWise export:
    Ledger,Category,Subcategory,Currency,Price,Account,Recorder,Date,Time,Note,Transaction
    """
    lines = ["Ledger,Category,Subcategory,Currency,Price,Account,Recorder,Date,Time,Note,Transaction"]
    
    for txn in transactions:
        if not include_invalid and not txn.is_valid:
            continue
        
        # Escape fields for CSV
        def escape_csv(val):
            if not val:
                return ""
            val = str(val)
            if "," in val or '"' in val or "\n" in val:
                return '"' + val.replace('"', '""') + '"'
            return val
        
        note = escape_csv(txn.description)
        category = escape_csv(txn.category) if txn.category else ""
        subcategory = escape_csv(txn.subcategory) if txn.subcategory else ""
        account = escape_csv(txn.account) if txn.account else "Paytm"
        
        line = f",{category},{subcategory},INR,{txn.amount or 0},{account},Import,{txn.date or ''},{txn.time or ''},{note},{txn.transaction_type}"
        lines.append(line)
    
    return "\n".join(lines)
